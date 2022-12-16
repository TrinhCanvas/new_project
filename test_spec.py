import time, json
from config import case_test
from test_flow import PagesTest
import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime
class Process_test(PagesTest):
	def __init__(self, config_channel):
		super().__init__(config_channel)
		# self.case_test = case_test
	# def file_result(self,name_page, status, msg, i):
	# 	workbook = xlsxwriter.Workbook('result.xlsx')
	# 	worksheet = workbook.add_worksheet()
	# 	worksheet.write('A1', 'NAME_PAGE')
	# 	worksheet.write('B1', 'STATUS')
	# 	worksheet.write('C1', 'ERROR')
	# 	row = 1+i
	# 	column = 0
	# 	for item in name_page:
	# 		worksheet.write(row, column, item)
	# 		column += 1
	# 		worksheet.write(row, column, status)
	# 		column += 1
	# 		worksheet.write(row, column, msg)
	# 		column = 0
	# 		row += 1
	# 	workbook.close()

	def file_data(self, name_page, status, msg, j):
		path = os.getcwd()
		file_name = r"\data.xlsx"
		if j == 0:
			wb = openpyxl.Workbook()
		else:
			wb = load_workbook(r'data.xlsx')
		sheet = wb.active
		content = [name_page, status, msg]
		i = 0
		for item in content:
			c1 = sheet.cell(row=1 + j, column=1 + i)
			c1.value = item
			wb.save(path + file_name)
			i += 1
	def create_result(self):
		wb2 = openpyxl.Workbook()
		ws2 = wb2.active
		wb1 = load_workbook(r"data.xlsx")
		ws1 = wb1.worksheets[0]
		mr = ws1.max_row
		mc = ws1.max_column

		for i in range(1, mr + 1):
			for j in range(1, mc + 1):
				c = ws1.cell(row=i, column=j)
				ws2.cell(row=i, column=j).value = c.value
		wb2.save(r"result.xlsx")

	def add_result(self, domain):
		wb1 = load_workbook(r"data.xlsx")
		ws1 = wb1.worksheets[0]
		wb2 = load_workbook(r"result.xlsx")
		ws2 = wb2.create_sheet(domain)
		mr = ws1.max_row
		mc = ws1.max_column

		for i in range(1, mr + 1):
			for j in range(1, mc + 1):
				c = ws1.cell(row=i, column=j)
				ws2.cell(row=i, column = j).value = c.value
		wb2.save(r"result.xlsx")

	def run_by_page(self):
		step_func = self.config_channel["step"]
		print("=========step_func=========")
		lst_start_time = {}
		final_detail_check = []
		flag_res_case = True
		msg = "Success"
		j=0
		for item in step_func:
			print(f"====start case: {item['name_func']}")
			if not flag_res_case:
				msg = "Compare case fail"
				# self.screenshot()
				if flag_res_case is None:
					break
			if "mobile" in item:
				if item["mobile"] == False:
					print("không test trên mobile")
					break
			if "name_time" in item and item["name_time"].startswith("start"):
				lst_start_time[item["name_time"].replace("start", "end")] = time.time()
			run_func = getattr(self, item["name_func"])
			result_case = run_func()
			# ===========================
			print("-------result_case-----------", result_case)
			self.file_data(self._driver.current_url, result_case, "check {} page {}".format(item["name_func"], self.config_channel["name_page"]), j)
			j += 1
			if result_case is False:
				self.send_msg_telegram("lỗi {} page {}".format(item["name_func"], self.config_channel["name_page"]))
				self.screenshot()
				self.send_img_telegram()
				pass
			# if result_case is None:
			# 	self.send_msg_telegram("lỗi {} page {}".format(item["name_func"], self.config_channel["name_page"]))
			# 	print('====case run error====')
			# 	self.screenshot()
			# 	self.send_img_telegram()
			# 	flag_res_case = False
			# 	msg = "Run case error internet"
			# 	self.file_result("False", "lỗi {} page {}".format(item["name_func"], self.config_channel["name_page"]))
			# 	break
			else:
				if not result_case:
					msg = "Fail"
				flag_res_case = result_case
			detail_case = {}
			if "name_case" in item:
				detail_case = {
					"name_case": item["name_case"],
					"status": result_case
				}
			if "name_time" in item and item["name_time"].startswith("end"):
				if item["name_time"] in lst_start_time:
					end_time = time.time()
					time_final = end_time - int(lst_start_time[item["name_time"]])
					flag_time = True
					if time_final > item["time_check"]:
						flag_time = False
						pass
					detail_case["time_check"] = {
						"time_run": time_final,
						"time_minimum": item["time_check"],
						"status": flag_time
					}
			if detail_case:
				final_detail_check.append(detail_case)
			if "time_sleep" in item:
				time.sleep(item["time_sleep"])

		response = {
			"name_case": self.config_channel["name_case"],
			"status": flag_res_case,
			"msg": msg,
			"detail_check": final_detail_check,
			"image_error": "/static/screenshot.png" if not flag_res_case else None
		}
		print(response)
		self.close_driver()